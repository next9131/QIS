import pickle as pk
import glob
import os
import pickle as pk
import numpy  as np
from   scipy import stats
#from   pandas.stats.api import ols
from   openpyxl import load_workbook


#if (__name__ == "__main__"):  # Execute when invoked from command line

def removeloadeddata(SaveFile):
#    print('Deleting *.dat files from current directory ...')
#    FileList = glob.glob('*.dat')
#    for file in FileList:
#        os.remove(file)
        
    print('Deleting ' + SaveFile + ' file from current directory ...')
    FileList = glob.glob(SaveFile)
    for file in FileList:
        os.remove(file)
    
    return 0

def initialize(): 
    print('Initializing data dictionary D ...')
    D               = {}
    return D

def loaddataFormat1(D,Filename):
    
    print('Loading spreadsheet ' + Filename + ' to dictionary D ...')
    wb                  = load_workbook(Filename) 
    asset               = wb.get_sheet_by_name('asset') 
    D['Asset']             = list()
    D['Sector_Asset']      = list()
    D['Industry_Asset']    = list()
    D['NumAsset']          = 0
    Var = 'FrogsHopping'
    while Var is not None:
        D['NumAsset']     = D['NumAsset'] + 1
        Var             = asset.cell(row = D['NumAsset']+1, column = 1).value
        SectorCode      = asset.cell(row = D['NumAsset']+1, column = 2).value
        IndustryCode    = asset.cell(row = D['NumAsset']+1, column = 3).value
        if Var is not None:               
            D['Asset'].append(Var)
            D['Sector_Asset'].append(SectorCode)
            D['Industry_Asset'].append(IndustryCode)
    D['NumAsset'] = D['NumAsset']-1   

    sector              = wb.get_sheet_by_name('sector') 
    D['Sector']         = list()
    D['SectorCode']     = list()
    D['NumSector']      = 0
    Var                 = 'FrogsHopping'
    while Var is not None:
        D['NumSector']     = D['NumSector'] + 1
        Var             = sector.cell(row = D['NumSector']+1, column = 1).value
        SectorCode      = sector.cell(row = D['NumSector']+1, column = 2).value
        if Var is not None:               
            D['Sector'].append(Var)
            D['SectorCode'].append(SectorCode)
    D['NumSector'] = D['NumSector']-1   

    industry        = wb.get_sheet_by_name('industry') 
    D['Industry']      = list()
    D['IndustryCode']  = list()
    D['NumIndustry']   = 0
    Var             = 'FrogsHopping'
    while Var is not None:
        D['NumIndustry']   = D['NumIndustry'] + 1
        Var             = industry.cell(row = D['NumIndustry']+1, column = 1).value
        IndustryCode    = industry.cell(row = D['NumIndustry']+1, column = 2).value
        if Var is not None:               
            D['Industry'].append(Var)
            D['IndustryCode'].append(IndustryCode)
    D['NumIndustry'] = D['NumIndustry']-1
    
    data            = wb.get_sheet_by_name('data') 
    D['Variable']      = set()
    Var             = 'FrogsHopping'
    D['dataNumRows']   = 1
    VarColNum       = 1
    while Var is not None:
        D['dataNumRows'] = D['dataNumRows'] + 1
        Var = data.cell(row = D['dataNumRows'], column = VarColNum).value
        if Var is not None:               
            D['Variable'].add(Var)
    D['dataNumRows']   =  D['dataNumRows']-1
    D['NumVariable']   = len(D['Variable'])      

    D['Period']        = list()
    D['NumPeriod']     = 0
    Var = 'FrogsHopping'
    while Var is not None:
        D['NumPeriod'] = D['NumPeriod'] + 1
        Var = data.cell(row = 1, column = D['NumPeriod']+2).value
        if Var is not None:               
            D['Period'].append(Var)
    D['NumPeriod'] = D['NumPeriod']-1   
    
    NumAsset    =   D['NumAsset']
    NumPeriod   =   D['NumPeriod']
    NAN_AP      =   np.nan*np.zeros((NumAsset,NumPeriod))
    for Var in D['Variable']:
        exec('D[' + "'" + Var + '_AP' + "'" +']'  + ' = NAN_AP.copy()')
#        exec('D.' + Var + '_AP = np.nan*np.zeros((D['NumAsset'], D['NumPeriod']))')
    print('Loading variables by asset and period from sheet data ...')    
    Var             = 'FrogsHopping'
    dataNumRows     = 1
    varColNum       = 1
    assetColNum     = 2
    while Var is not None:
        dataNumRows = dataNumRows + 1
        Var     = data.cell(row = dataNumRows, column = varColNum).value
        asset   = data.cell(row = dataNumRows, column = assetColNum).value
#        print(Var)
#        print([dataNumRows,varColNum,assetColNum])
        if Var is not None:
            assetpos    =  D['Asset'].index(asset)
            for period in D['Period']:
                periodpos    =  D['Period'].index(period)
                Number  = data.cell(row = dataNumRows, column = periodpos+3).value
#                print([assetpos,periodpos,Number])
#                print('D[' + "'" + Var + '_AP' + "'" +']'  + '[assetpos,periodpos] = Number')
                exec('D[' + "'" + Var + '_AP' + "'" +']'  + '[assetpos,periodpos] = Number')
#                print('D.'+Var+'_AP[assetpos,periodpos] = Number')
    print('Variables by asset and period from sheet data are loaded.')             
#    print(D.AnalystAgreementRevisions_AP[1,1])
    print('Number of rows in sheet data: ' + str(D['dataNumRows']))
    print('            Number of Assets: ' + str(D['NumAsset']))
    print('           Number of Periods: ' + str(D['NumPeriod']))
    print('           Number of Sectors: ' + str(D['NumSector']))
    print('         Number of Industrys: ' + str(D['NumIndustry']))
    print('         Number of Variables: ' + str(D['NumVariable']))
    ForwardReturns_AP   = NAN_AP.copy()
    Returns_AP          = NAN_AP.copy()
    for p in range(1,D['NumPeriod']):
        for a in range(0,D['NumAsset']):
            Returns_AP[a,p] = np.log(D['Price_AP'][a,p]/D['Price_AP'][a,p-1])
    for p in range(0,D['NumPeriod']-1):
        for a in range(0,D['NumAsset']):
            ForwardReturns_AP[a,p] = np.log(D['Price_AP'][a,p+1]/D['Price_AP'][a,p])        
    D['ForwardReturns_AP']  = ForwardReturns_AP
    D['Returns_AP']         = Returns_AP
    return D

def saveproject(D,SaveFile):
    print('Saving D to ' + SaveFile)
    handle  = open(SaveFile, 'wb')
    pk.dump(D,handle)
    print('D has been saved to ' + SaveFile)
    return 1

def loadproject(SaveFile):
    print('Loading from ' + SaveFile)
    file    = open(SaveFile, 'rb')
    D       = pk.load(file)
    print('Data loaded from ' + SaveFile)
    return D


def savefile(SaveFile,X):
    print('Saving to ' + SaveFile)
    handle  = open(SaveFile, 'wb')
    pk.dump(X,handle)
    print('saved to ' + SaveFile)
    return 1

def loadsavedfile(SaveFile):
    print('Loading from ' + SaveFile)
    file    = open(SaveFile, 'rb')
    F       = pk.load(file)
    print('Loaded from ' + SaveFile)
    return F